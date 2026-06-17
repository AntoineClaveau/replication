import streamlit as st
import pandas as pd
import numpy as np
from sklearn.neighbors import BallTree
import folium
from folium.plugins import PolyLineTextPath, HeatMap, MarkerCluster
from streamlit_folium import st_folium
import os
import matplotlib.pyplot as plt
import plotly.express as px
from scipy.ndimage import gaussian_filter
import io
import zipfile
from dataclasses import dataclass
from datetime import datetime
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import geopandas as gpd
import contextily as ctx
from shapely.geometry import Point
from matplotlib.patches import FancyBboxPatch, Circle
import matplotlib.image as mpimg

# ======================================================
# CONFIG
# ======================================================
KMS_PER_RADIAN = 6371.0088
try:
    # Ce chemin est spécifique à un poste Windows local (dev).
    # Sur Streamlit Community Cloud (Linux) ce dossier n'existe pas :
    # on ignore simplement l'erreur et on reste au répertoire du repo,
    # où EPRTR.csv / UWWTPS.csv doivent être placés (à côté de app.py).
    os.chdir("d:/synologydrive/wp2/t2.3/App")
except (FileNotFoundError, NotADirectoryError, OSError):
    pass

st.set_page_config(layout="wide")
st.title("Industrial symbiosis replication analysis")

import base64

def _safe_b64_image(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        return None

logo_b64 = _safe_b64_image("Strane-logo.png")

if logo_b64:
    st.markdown(f"""
    <style>
    [data-testid="stSidebar"] {{
        background-color: #CEFAD6;
        background-image: url("data:image/png;base64,{logo_b64}");
        background-repeat: no-repeat;
        background-position: right bottom;
        background-size: 120px auto;
        padding-bottom: 140px;
    }}
    </style>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { background-color: #CEFAD6; }
    </style>
    """, unsafe_allow_html=True)

# ======================================================
# DATA STRUCTURE
# ======================================================
@dataclass
class ExportArtifacts:
    summary_df: pd.DataFrame
    networks_df: pd.DataFrame
    actors_df: pd.DataFrame
    links_df: pd.DataFrame
    country_stats_df: pd.DataFrame
    parameters_df: pd.DataFrame
    readme_df: pd.DataFrame


# ======================================================
# SESSION STATE
# ======================================================
if "paths" not in st.session_state:
    st.session_state.paths = []
if "centroids_df" not in st.session_state:
    st.session_state.centroids_df = None
if "selected_path" not in st.session_state:
    st.session_state.selected_path = None
if "layer_defs" not in st.session_state:
    st.session_state.layer_defs = [
        {
            "label": "Emitter A (EPRTR)",
            "source": "EPRTR",
            "logic": "AND",
            "actor_groups": [
                {
                    "label": "Huilerie",
                    "codes": ["10.41"],
                    "min_actors": 1,
                },
            ],
            "max_to_next_km": 30.0,
        },
        {
            "label": "Intermediate (UWWTPS)",
            "source": "UWWTPS",
            "codes": "",
            "cap_min": 0.0,
            "cap_max": 30000.0,
            "min_actors": 1,
            "max_to_next_km": 50.0,
        },
        {
            "label": "End-user (EPRTR)",
            "source": "EPRTR",
            "codes": "20.16",
            "min_actors": 1,
            "max_to_next_km": None,
        },
    ]

if "eprtr_df" not in st.session_state:
    st.session_state.eprtr_df = None
if "uww_df" not in st.session_state:
    st.session_state.uww_df = None
if "layer_dfs" not in st.session_state:
    st.session_state.layer_dfs = None

if "pivot_layer" not in st.session_state:
    st.session_state.pivot_layer = 1
if "pivot_mode" not in st.session_state:
    st.session_state.pivot_mode = "Pivot stakeholder (selection)"
if "pivot_selected_idx" not in st.session_state:
    st.session_state.pivot_selected_idx = 0
if "heatmap_data" not in st.session_state:
    st.session_state.heatmap_data = None

if "export_html" not in st.session_state:
    st.session_state.export_html = None
if "export_excel" not in st.session_state:
    st.session_state.export_excel = None
if "export_pdf" not in st.session_state:
    st.session_state.export_pdf = None
if "export_zip" not in st.session_state:
    st.session_state.export_zip = None

if "export_artifacts" not in st.session_state:
    st.session_state.export_artifacts = None


# ======================================================
# UTILS
# ======================================================
def clean_cols(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip().str.lower()
    return df


def clean_latlon(df: pd.DataFrame) -> pd.DataFrame:
    df["latitude"] = pd.to_numeric(df["latitude"], errors="coerce")
    df["longitude"] = pd.to_numeric(df["longitude"], errors="coerce")
    return df.dropna(subset=["latitude", "longitude"]).copy()


@st.cache_data(show_spinner=False)
def load_eprtr(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", encoding="latin1", low_memory=False)
    return clean_latlon(clean_cols(df))


@st.cache_data(show_spinner=False)
def load_uwwtps(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", low_memory=False)
    df = clean_latlon(clean_cols(df))
    if "uwwcapacity" in df.columns:
        df["uwwcapacity"] = pd.to_numeric(df["uwwcapacity"], errors="coerce")
    return df


def make_base_map(lat, lon, zoom=6):
    return folium.Map(location=[lat, lon], zoom_start=zoom, tiles="CartoDB positron")


def parse_codes(txt: str):
    if txt is None:
        return []
    parts = [p.strip() for p in str(txt).replace(";", ",").split(",")]
    return [p for p in parts if p]


def nearest_centroid_id(click_lat, click_lon, centroids_df, max_km=5.0):
    if centroids_df is None or len(centroids_df) == 0:
        return None
    tree = BallTree(np.radians(centroids_df[["lat", "lon"]].values), metric="haversine")
    dist, ind = tree.query(np.radians([[click_lat, click_lon]]), k=1)
    pid = int(centroids_df.iloc[int(ind[0][0])]["path_id"])
    d_km = float(dist[0][0]) * KMS_PER_RADIAN
    if d_km > max_km:
        return None
    return pid


def get_actor_country(row):
    src = row.get("__layer_source", "")
    if src == "EPRTR":
        country = row.get("countrycode", "")
    elif src == "UWWTPS":
        country = row.get("rptmstatekey", row.get("countrycode", ""))
    else:
        country = row.get("countrycode", "")
    if pd.isna(country) or str(country).strip() == "":
        return "NA"
    return str(country).upper()


def haversine_km(lat1, lon1, lat2, lon2):
    lat1_r, lon1_r, lat2_r, lon2_r = map(np.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2_r - lat1_r
    dlon = lon2_r - lon1_r
    a = np.sin(dlat / 2.0) ** 2 + np.cos(lat1_r) * np.cos(lat2_r) * np.sin(dlon / 2.0) ** 2
    return 2.0 * KMS_PER_RADIAN * np.arcsin(np.sqrt(a))


def filter_layer_df(layer_def: dict, eprtr: pd.DataFrame, uww: pd.DataFrame) -> pd.DataFrame:
    src = layer_def["source"]

    if src == "EPRTR":
        df = eprtr.copy()
        all_codes = [
            c.strip()
            for g in layer_def.get("actor_groups", [])
            for c in g.get("codes", [])
            if c and str(c).strip()
        ]

        if all_codes:
            naf_series = df["nacemaineconomicactivitycode"].astype(str)
            mask = pd.Series(False, index=df.index)
            for code in all_codes:
                mask = mask | naf_series.str.startswith(code)
            df = df[mask].copy()

        df = df.reset_index(drop=True)
        df["__layer_source"] = "EPRTR"
        df["__layer_code"] = df["nacemaineconomicactivitycode"].astype(str)
        df["__name"] = df.get("facilityname", pd.Series([""] * len(df))).astype(str)
        return df

    if src == "UWWTPS":
        df = uww.copy()
        cap_min = float(layer_def.get("cap_min", 0.0))
        cap_max = float(layer_def.get("cap_max", 1e12))

        if "uwwcapacity" in df.columns:
            df["uwwcapacity"] = pd.to_numeric(df["uwwcapacity"], errors="coerce")
            df = df[
                (df["uwwcapacity"] >= cap_min) &
                (df["uwwcapacity"] <= cap_max)
            ].copy()

        df = df.reset_index(drop=True)
        df["__layer_source"] = "UWWTPS"
        df["__layer_code"] = "UWWTPS"
        if "uwwname" in df.columns:
            df["__name"] = df["uwwname"].astype(str)
        else:
            df["__name"] = pd.Series(["UWWTPS"] * len(df)).astype(str)
        return df

    return pd.DataFrame()


def _build_tree(df: pd.DataFrame) -> BallTree:
    coords = np.radians(df[["latitude", "longitude"]].values)
    return BallTree(coords, metric="haversine")


def nearest_actor_idx(tree: BallTree, df: pd.DataFrame, lat: float, lon: float):
    dist, ind = tree.query(np.radians([[lat, lon]]), k=1)
    return int(ind[0][0])


def _query_within_km(tree: BallTree, center_lat, center_lon, radius_km: float, max_keep: int | None):
    r = float(radius_km) / KMS_PER_RADIAN
    center = np.radians([[center_lat, center_lon]])
    idxs, dists = tree.query_radius(center, r=r, return_distance=True)
    idxs = idxs[0]
    dists = dists[0]
    if len(idxs) == 0:
        return []
    order = np.argsort(dists)
    idxs = idxs[order]
    if max_keep is not None and len(idxs) > int(max_keep):
        idxs = idxs[: int(max_keep)]
    return [int(x) for x in idxs.tolist()]


def _labels_for_layer(df: pd.DataFrame):
    labels = []
    for irow, r in df.iterrows():
        nm = str(r.get("__name", ""))
        cd = str(r.get("__layer_code", ""))
        cc = str(get_actor_country(r))
        labels.append(f"[{irow}] {cd} • {cc} • {nm}"[:200])
    return labels


def _pivot_radius_km(layer_defs, pivot_layer: int):
    cand = []
    if pivot_layer - 1 >= 0:
        d = layer_defs[pivot_layer - 1].get("max_to_next_km", None)
        if d is not None:
            cand.append(float(d))
    if pivot_layer < len(layer_defs) - 1:
        d = layer_defs[pivot_layer].get("max_to_next_km", None)
        if d is not None:
            cand.append(float(d))
    return max(cand) if cand else 99999.0


def validate_actor_groups(df, idxs, layer_def):
    if layer_def.get("source") == "UWWTPS":
        return True

    logic = layer_def.get("logic", "AND")
    groups = layer_def.get("actor_groups", [])

    if not groups:
        return True

    layer_codes = df.iloc[idxs]["__layer_code"].astype(str)
    results = []

    for g in groups:
        codes = [c for c in g.get("codes", []) if c]
        min_req = int(g.get("min_actors", 1))

        count = 0
        for c in codes:
            count += layer_codes.str.startswith(c).sum()

        results.append(count >= min_req)

    return all(results) if logic == "AND" else any(results)


def build_networks_bidirectional_pivot(
    layer_defs,
    layer_dfs,
    pivot_layer: int,
    pivot_indices: list[int],
    max_neighbors=200,
    max_networks=20000,
    progress_callback=None,
):
    n_layers = len(layer_defs)
    if n_layers < 2:
        return []

    for i in range(n_layers):
        if layer_dfs[i].empty:
            return []

    trees = [_build_tree(layer_dfs[i]) for i in range(n_layers)]

    out = []
    pivot_layer = int(pivot_layer)
    n_pivots = len(pivot_indices)

    for step, piv_idx in enumerate(pivot_indices):
        if progress_callback is not None:
            progress_callback(step, n_pivots)

        dfP = layer_dfs[pivot_layer]
        if piv_idx < 0 or piv_idx >= len(dfP):
            continue

        piv_row = dfP.iloc[int(piv_idx)]
        piv_lat = float(piv_row.latitude)
        piv_lon = float(piv_row.longitude)

        min_p = int(layer_defs[pivot_layer].get("min_actors", 1))
        if min_p <= 1:
            pivot_set = [int(piv_idx)]
        else:
            pr_km = _pivot_radius_km(layer_defs, pivot_layer)
            pivot_set = _query_within_km(
                trees[pivot_layer],
                piv_lat, piv_lon,
                radius_km=pr_km,
                max_keep=max_neighbors
            )
            if int(piv_idx) not in pivot_set:
                pivot_set = [int(piv_idx)] + pivot_set
            pivot_set = list(dict.fromkeys(pivot_set))
            if len(pivot_set) < min_p:
                continue
            if len(pivot_set) > max_neighbors:
                pivot_set = pivot_set[:max_neighbors]

        dfP_sel = dfP.iloc[pivot_set]
        piv_c_lat = float(dfP_sel.latitude.mean())
        piv_c_lon = float(dfP_sel.longitude.mean())

        layer_sets = [None] * n_layers
        layer_centroids = [None] * n_layers

        layer_sets[pivot_layer] = pivot_set
        layer_centroids[pivot_layer] = (piv_c_lat, piv_c_lon)

        ok = True

        for i in range(pivot_layer, n_layers - 1):
            d_km = layer_defs[i].get("max_to_next_km", None)
            if d_km is None:
                ok = False
                break

            center_lat, center_lon = layer_centroids[i]
            idxs_next = _query_within_km(
                trees[i + 1],
                center_lat, center_lon,
                radius_km=float(d_km),
                max_keep=max_neighbors
            )

            if not validate_actor_groups(layer_dfs[i + 1], idxs_next, layer_defs[i + 1]):
                ok = False
                break

            layer_sets[i + 1] = idxs_next
            df_next = layer_dfs[i + 1].iloc[idxs_next]
            layer_centroids[i + 1] = (
                float(df_next.latitude.mean()),
                float(df_next.longitude.mean())
            )

        if not ok:
            continue

        for i in range(pivot_layer - 1, -1, -1):
            d_km = layer_defs[i].get("max_to_next_km", None)
            if d_km is None:
                ok = False
                break

            center_lat, center_lon = layer_centroids[i + 1]
            idxs_i = _query_within_km(
                trees[i],
                center_lat, center_lon,
                radius_km=float(d_km),
                max_keep=max_neighbors
            )

            if not validate_actor_groups(layer_dfs[i], idxs_i, layer_defs[i]):
                ok = False
                break

            layer_sets[i] = idxs_i
            df_i = layer_dfs[i].iloc[idxs_i]
            layer_centroids[i] = (
                float(df_i.latitude.mean()),
                float(df_i.longitude.mean())
            )

        if not ok:
            continue

        lats = [c[0] for c in layer_centroids if c is not None]
        lons = [c[1] for c in layer_centroids if c is not None]
        if not lats:
            continue

        out.append({
            "path_id": len(out),
            "pivot_layer": pivot_layer,
            "pivot_idx": int(piv_idx),
            "layer_sets": layer_sets,
            "layer_centroids": layer_centroids,
            "lat": float(np.mean(lats)),
            "lon": float(np.mean(lons)),
        })

        if len(out) >= int(max_networks):
            break

    return out

def centroids_to_gdf(centroids_df):
    if centroids_df is None or len(centroids_df) == 0:
        return None

    gdf = gpd.GeoDataFrame(
        centroids_df.copy(),
        geometry=gpd.points_from_xy(centroids_df["lon"], centroids_df["lat"]),
        crs="EPSG:4326"
    )
    return gdf.to_crs(epsg=3857)

def draw_centroid_map_page(pdf, centroids_df):
    fig = plt.figure(figsize=(11.69, 8.27))
    ax = fig.add_axes([0.05, 0.08, 0.82, 0.78])

    draw_report_header(
        fig,
        "Detected network centroids",
        "Geographical distribution of detected replication network centroids"
    )

    gdf = centroids_to_gdf(centroids_df)

    if gdf is None or gdf.empty:
        ax.text(0.5, 0.5, "No centroid data available", ha="center", va="center", transform=ax.transAxes)
        ax.axis("off")
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)
        return

    xmin, ymin, xmax, ymax = gdf.total_bounds
    dx = max((xmax - xmin) * 0.15, 20000)
    dy = max((ymax - ymin) * 0.15, 20000)

    xmin -= dx
    xmax += dx
    ymin -= dy
    ymax += dy

    gdf.plot(
        ax=ax,
        markersize=45,
        color="#00B400",
        edgecolor="black",
        linewidth=0.5,
        alpha=0.9,
        label="Detected network centroid",
        zorder=3
    )

    ax.set_xlim(xmin, xmax)
    ax.set_ylim(ymin, ymax)

    ctx.add_basemap(
        ax,
        source=ctx.providers.OpenStreetMap.Mapnik,
        crs=gdf.crs,
        reset_extent=False,
        attribution_size=6
    )

    ax.set_xlim(xmin, xmax)
    ax.set_ylim(ymin, ymax)

    ax.set_axis_off()
    ax.legend(loc="upper right", frameon=True)

    fig.text(
        0.08, 0.06,
        "Each point represents the centroid of one detected multi-stage industrial symbiosis network.",
        fontsize=10,
        color="#555555"
    )

    pdf.savefig(fig, bbox_inches="tight", dpi=200)
    plt.close(fig)


def draw_centroid_heatmap_page(pdf, centroids_df):
    from matplotlib.colors import ListedColormap

    fig = plt.figure(figsize=(11.69, 8.27))
    ax = fig.add_axes([0.05, 0.08, 0.82, 0.78])

    draw_report_header(
        fig,
        "Heatmap of detected network centroids",
        "Spatial concentration of detected replication opportunities"
    )

    gdf = centroids_to_gdf(centroids_df)

    if gdf is None or gdf.empty:
        ax.text(0.5, 0.5, "No centroid data available", ha="center", va="center", transform=ax.transAxes)
        ax.axis("off")
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)
        return

    x = gdf.geometry.x.values
    y = gdf.geometry.y.values

    # Emprise basée sur les points détectés
    xmin, ymin, xmax, ymax = gdf.total_bounds
    dx = max((xmax - xmin) * 0.15, 20000)
    dy = max((ymax - ymin) * 0.15, 20000)

    xmin -= dx
    xmax += dx
    ymin -= dy
    ymax += dy

    # Heatmap raster
    bins = 180
    heat, xedges, yedges = np.histogram2d(
        x, y,
        bins=bins,
        range=[[xmin, xmax], [ymin, ymax]]
    )

    # Lissage : un flou gaussien équivalent au lissage manuel précédent
    # (4 passes de moyenne 3x3 via np.roll), mais calculé en une seule
    # opération vectorisée par scipy — même rendu visuel, plus rapide.
    # mode="nearest" évite l'effet de bord "torique" du np.roll original
    # (sans incidence visible : les bords du raster sont à densité quasi nulle).
    heat = gaussian_filter(heat, sigma=1.6, mode="nearest")

    # Normalisation
    if heat.max() > 0:
        heat_norm = heat / heat.max()
    else:
        heat_norm = heat.copy()

    # Colormap avec transparence sur faibles valeurs
    base = plt.get_cmap("hot", 256)
    colors = base(np.linspace(0, 1, 256))

    alphas = np.zeros(256)
    for i in range(256):
        v = i / 255.0
        if v < 0.10:
            alphas[i] = 0.0
        elif v < 0.25:
            alphas[i] = (v - 0.10) / (0.25 - 0.10) * 0.35
        else:
            alphas[i] = 0.35 + (v - 0.25) / 0.75 * 0.55

    colors[:, -1] = np.clip(alphas, 0, 0.9)
    transparent_hot = ListedColormap(colors)

    extent = [xedges[0], xedges[-1], yedges[0], yedges[-1]]

    # 1) on dessine d'abord la heatmap
    img = ax.imshow(
        heat_norm.T,
        extent=extent,
        origin="lower",
        cmap=transparent_hot,
        interpolation="bilinear",
        zorder=3
    )

    # 2) on fixe l'emprise
    ax.set_xlim(xmin, xmax)
    ax.set_ylim(ymin, ymax)

    # 3) on ajoute le fond OSM sans réinitialiser l'emprise
    ctx.add_basemap(
        ax,
        source=ctx.providers.OpenStreetMap.Mapnik,
        crs=gdf.crs,
        reset_extent=False,
        attribution_size=6
    )

    # 4) on réapplique l'emprise par sécurité
    ax.set_xlim(xmin, xmax)
    ax.set_ylim(ymin, ymax)

    cbar = fig.colorbar(img, ax=ax, shrink=0.8)
    cbar.set_label("Relative centroid density")

    ax.set_axis_off()

    fig.text(
        0.08, 0.06,
        "The heatmap highlights areas with a higher concentration of detected network centroids. "
        "Lower-density areas are displayed with higher transparency.",
        fontsize=10,
        color="#555555"
    )

    pdf.savefig(fig, bbox_inches="tight", dpi=200)
    plt.close(fig)



def get_europe_extent_3857():
    """
    Approximate Europe extent in EPSG:3857
    lon_min, lat_min, lon_max, lat_max in WGS84:
    [-12, 35] to [32, 72]
    """
    europe_box = gpd.GeoSeries(
        [Point(-12, 35), Point(32, 72)],
        crs="EPSG:4326"
    ).to_crs(epsg=3857)

    xmin = europe_box.iloc[0].x
    ymin = europe_box.iloc[0].y
    xmax = europe_box.iloc[1].x
    ymax = europe_box.iloc[1].y
    return xmin, ymin, xmax, ymax


def _stage_fill_color(li, src, n_layers):
    if src == "UWWTPS":
        return "#FCE5CD"  # orange clair
    if li == 0:
        return "#F4CCCC"  # rouge clair
    if li == n_layers - 1:
        return "#CFE2F3"  # bleu clair
    return "#EADCF8"      # violet clair


def _wrap_codes(codes):
    if not codes:
        return "-"
    if isinstance(codes, str):
        return codes
    return ", ".join([str(c) for c in codes if str(c).strip()])


def _draw_box(ax, x, y, w, h, text, fc="#FFFFFF", ec="#444444", fontsize=9, lw=1.2, rounded=0.02, weight="normal"):
    patch = FancyBboxPatch(
        (x, y), w, h,
        boxstyle=f"round,pad=0.01,rounding_size={rounded}",
        linewidth=lw,
        edgecolor=ec,
        facecolor=fc
    )
    ax.add_patch(patch)
    ax.text(
        x + w / 2,
        y + h / 2,
        text,
        ha="center",
        va="center",
        fontsize=fontsize,
        weight=weight,
        wrap=True
    )
    return patch


def draw_methodology_flow_page(pdf, layer_defs):
    fig = plt.figure(figsize=(11.69, 8.27))
    ax = fig.add_axes([0.04, 0.10, 0.92, 0.80])
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")

    draw_report_header(
        fig,
        "Methodology flow diagram",
        "Visual summary of the detection model configuration"
    )

    n_layers = len(layer_defs)
    if n_layers == 0:
        ax.text(0.5, 0.5, "No stage configuration available", ha="center", va="center", fontsize=14)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)
        return

    margin_x = 0.03
    gap_x = 0.035
    stage_w = (1 - 2 * margin_x - (n_layers - 1) * gap_x) / n_layers
    stage_top = 0.88
    stage_bottom = 0.12
    stage_h = stage_top - stage_bottom

    for i, ld in enumerate(layer_defs):
        x0 = margin_x + i * (stage_w + gap_x)
        y0 = stage_bottom
        src = ld.get("source", "")
        logic = ld.get("logic", "AND")
        label = ld.get("label", f"Stage {i+1}")
        fc = _stage_fill_color(i, src, n_layers)

        # Outer stage panel
        _draw_box(
            ax, x0, y0, stage_w, stage_h,
            "",
            fc="#FAFAFA",
            ec="#999999",
            lw=1.0,
            rounded=0.015
        )

        # Stage header
        _draw_box(
            ax, x0 + 0.01, stage_top - 0.09, stage_w - 0.02, 0.08,
            f"Stage {i+1}\n{label}\nSource: {src}",
            fc=fc,
            ec="#666666",
            fontsize=10,
            lw=1.2,
            rounded=0.015,
            weight="bold"
        )

        # Logic box
        logic_text = f"Logic: {logic}" if src == "EPRTR" else "Logic: n/a"
        _draw_box(
            ax, x0 + 0.04, stage_top - 0.16, stage_w - 0.08, 0.045,
            logic_text,
            fc="#EFEFEF",
            ec="#777777",
            fontsize=9,
            lw=1.0,
            rounded=0.01,
            weight="bold"
        )

        # Actor/group boxes
        box_y = stage_top - 0.24
        box_h = 0.075
        box_gap = 0.015

        if src == "UWWTPS":
            min_actors = 1
            actor_groups = ld.get("actor_groups", [])
            if actor_groups:
                min_actors = int(actor_groups[0].get("min_actors", 1))

            cap_min = ld.get("cap_min", 0)
            cap_max = ld.get("cap_max", "")
            n_repeat = max(1, min_actors)

            for r in range(n_repeat):
                txt = (
                    f"UWWTPS\n"
                    f"Capacity: {cap_min} to {cap_max}\n"
                    f"Min actors: {min_actors}"
                )
                _draw_box(
                    ax, x0 + 0.03, box_y - r * (box_h + box_gap), stage_w - 0.06, box_h,
                    txt,
                    fc="#FFF2CC",
                    ec="#A67C00",
                    fontsize=8.5,
                    lw=1.1,
                    rounded=0.012
                )
        else:
            groups = ld.get("actor_groups", [])
            current_y = box_y

            for g in groups:
                g_label = g.get("label", "Actor")
                codes = _wrap_codes(g.get("codes", []))
                min_actors = int(g.get("min_actors", 1))
                n_repeat = max(1, min_actors)

                for r in range(n_repeat):
                    txt = (
                        f"{g_label}\n"
                        f"NAF/NACE: {codes}\n"
                        f"Min actors: {min_actors}"
                    )
                    _draw_box(
                        ax, x0 + 0.03, current_y, stage_w - 0.06, box_h,
                        txt,
                        fc="#FFFFFF",
                        ec="#666666",
                        fontsize=8.3,
                        lw=1.0,
                        rounded=0.012
                    )
                    current_y -= (box_h + box_gap)

                current_y -= 0.012

        # Arrow to next stage
        if i < n_layers - 1:
            next_dist = ld.get("max_to_next_km", None)
            x_start = x0 + stage_w
            x_end = x0 + stage_w + gap_x
            y_mid = y0 + stage_h * 0.55

            ax.annotate(
                "",
                xy=(x_end - 0.005, y_mid),
                xytext=(x_start + 0.005, y_mid),
                arrowprops=dict(arrowstyle="->", lw=2, color="#444444")
            )

            dist_txt = f"Max distance\n{next_dist} km" if next_dist is not None else "No distance"
            _draw_box(
                ax,
                x_start + 0.005,
                y_mid + 0.025,
                gap_x - 0.01,
                0.07,
                dist_txt,
                fc="#F3F3F3",
                ec="#888888",
                fontsize=8,
                lw=0.9,
                rounded=0.01
            )

    fig.text(
        0.05, 0.04,
        "Actor boxes are repeated visually according to the minimum number of actors required for each group.",
        fontsize=10,
        color="#666666"
    )

    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def draw_example_network_page(pdf, paths, layer_dfs, layer_defs, network_id=0):
    fig = plt.figure(figsize=(11.69, 8.27))
    ax = fig.add_axes([0.04, 0.10, 0.92, 0.80])
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")

    draw_report_header(
        fig,
        "Example detected network diagram",
        "Illustrative actor-to-actor structure for one detected network"
    )

    if not paths:
        ax.text(0.5, 0.5, "No detected network available", ha="center", va="center", fontsize=14)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)
        return

    network_id = min(max(0, int(network_id)), len(paths) - 1)
    p = paths[network_id]
    layer_sets = p["layer_sets"]
    n_layers = len(layer_sets)

    margin_x = 0.03
    gap_x = 0.035
    stage_w = (1 - 2 * margin_x - (n_layers - 1) * gap_x) / n_layers
    stage_top = 0.88
    stage_bottom = 0.12
    stage_h = stage_top - stage_bottom

    actor_positions = {}

    # --------------------------------------------------
    # Draw stage panels + actor boxes
    # --------------------------------------------------
    for i, idxs in enumerate(layer_sets):
        x0 = margin_x + i * (stage_w + gap_x)
        y0 = stage_bottom
        ld = layer_defs[i]
        src = ld.get("source", "")
        label = ld.get("label", f"Stage {i+1}")
        fc = _stage_fill_color(i, src, n_layers)

        _draw_box(
            ax, x0, y0, stage_w, stage_h,
            "",
            fc="#FAFAFA",
            ec="#999999",
            lw=1.0,
            rounded=0.015
        )

        _draw_box(
            ax, x0 + 0.01, stage_top - 0.09, stage_w - 0.02, 0.08,
            f"Stage {i+1}\n{label}\nSource: {src}",
            fc=fc,
            ec="#666666",
            fontsize=10,
            lw=1.2,
            rounded=0.015,
            weight="bold"
        )

        if not idxs:
            ax.text(x0 + stage_w / 2, y0 + stage_h / 2, "No actor", ha="center", va="center", fontsize=10)
            actor_positions[i] = []
            continue

        dfL = layer_dfs[i]
        n_boxes = len(idxs)
        usable_top = stage_top - 0.14
        usable_bottom = stage_bottom + 0.04
        usable_h = usable_top - usable_bottom

        box_h = min(0.08, usable_h / max(n_boxes, 1) * 0.75)
        gap_y = min(0.02, usable_h / max(n_boxes, 2) * 0.25)

        total_h = n_boxes * box_h + (n_boxes - 1) * gap_y
        start_y = usable_bottom + (usable_h - total_h) / 2 + (n_boxes - 1) * (box_h + gap_y)

        stage_actor_positions = []

        for k, ridx in enumerate(idxs):
            row = dfL.iloc[int(ridx)]
            name = str(row.get("__name", "")).strip()
            country = get_actor_country(row)
            code = str(row.get("__layer_code", "")).strip()

            # pays entre parenthèses à côté du nom
            title = f"{name} ({country})" if country not in ["", "NA"] else name
            title = title[:55]

            txt = f"{title}\nNAF/NACE: {code}"

            y_box = start_y - k * (box_h + gap_y)
            _draw_box(
                ax,
                x0 + 0.03,
                y_box,
                stage_w - 0.06,
                box_h,
                txt,
                fc="#FFFFFF",
                ec="#555555",
                fontsize=8.0,
                lw=1.0,
                rounded=0.012
            )

            stage_actor_positions.append({
                "center_left": (x0 + 0.03, y_box + box_h / 2),
                "center_right": (x0 + stage_w - 0.03, y_box + box_h / 2),
                "anchor_in": (x0 + 0.03, y_box + box_h / 2),
                "anchor_out": (x0 + stage_w - 0.03, y_box + box_h / 2),
                "row": row,
                "box_y": y_box,
                "box_h": box_h,
            })

        actor_positions[i] = stage_actor_positions

    # --------------------------------------------------
    # Draw ALL actor-to-actor links between consecutive stages
    # with real distance labels
    # --------------------------------------------------
    for li in range(n_layers - 1):
        idxs_src = layer_sets[li]
        idxs_tgt = layer_sets[li + 1]

        if not idxs_src or not idxs_tgt:
            continue

        src_positions = actor_positions.get(li, [])
        tgt_positions = actor_positions.get(li + 1, [])

        if not src_positions or not tgt_positions:
            continue

        # For each actor in stage i, connect to ALL actors in stage i+1
        for isrc, src_info in enumerate(src_positions):
            src_row = src_info["row"]
            src_lat = float(src_row.latitude)
            src_lon = float(src_row.longitude)

            for jtgt, tgt_info in enumerate(tgt_positions):
                tgt_row = tgt_info["row"]
                tgt_lat = float(tgt_row.latitude)
                tgt_lon = float(tgt_row.longitude)

                dist_km = haversine_km(src_lat, src_lon, tgt_lat, tgt_lon)

                x1, y1 = src_info["anchor_out"]
                x2, y2 = tgt_info["anchor_in"]

                # léger décalage vertical pour éviter superposition totale
                n_tgt = max(len(tgt_positions), 1)
                offset_src = ((jtgt - (n_tgt - 1) / 2.0) / max(n_tgt, 2)) * 0.03
                y1_adj = y1 + offset_src
                y2_adj = y2

                ax.annotate(
                    "",
                    xy=(x2, y2_adj),
                    xytext=(x1, y1_adj),
                    arrowprops=dict(
                        arrowstyle="->",
                        lw=0.9,
                        color="#333333",
                        alpha=0.65,
                        shrinkA=2,
                        shrinkB=2
                    )
                )

                # position du label de distance
                xm = (x1 + x2) / 2.0
                ym = (y1_adj + y2_adj) / 2.0

                # petit fond blanc pour lisibilité
                ax.text(
                    xm,
                    ym,
                    f"{dist_km:.1f} km",
                    fontsize=6.8,
                    ha="center",
                    va="center",
                    color="#222222",
                    bbox=dict(
                        boxstyle="round,pad=0.15",
                        facecolor="white",
                        edgecolor="#BBBBBB",
                        linewidth=0.5,
                        alpha=0.85
                    )
                )

    fig.text(
        0.05, 0.04,
        f"Displayed example: detected network #{network_id}. "
        f"All actor-to-actor links between consecutive stages are shown, with real geographic distance labels.",
        fontsize=10,
        color="#666666"
    )

    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def draw_balltree_method_page(pdf):
    fig = plt.figure(figsize=(11.69, 8.27))
    ax = fig.add_axes([0.04, 0.10, 0.92, 0.78])
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")

    draw_report_header(
        fig,
        "How replication opportunities are detected?",
        "BallTree-based multi-stage spatial matching of industrial actors"
    )

    theme_color = "#02A717"

    def add_box(x, y, w, h, text, fc="#FFFFFF", ec="#666666", fontsize=10, weight="normal"):
        patch = FancyBboxPatch(
            (x, y), w, h,
            boxstyle="round,pad=0.012,rounding_size=0.015",
            linewidth=1.2,
            edgecolor=ec,
            facecolor=fc
        )
        ax.add_patch(patch)
        ax.text(
            x + w / 2, y + h / 2, text,
            ha="center", va="center",
            fontsize=fontsize, weight=weight, wrap=True
        )
        return patch

    def add_arrow(x1, y1, x2, y2, text=None):
        ax.annotate(
            "",
            xy=(x2, y2),
            xytext=(x1, y1),
            arrowprops=dict(arrowstyle="->", lw=2.0, color="#4D4D4D")
        )
        if text:
            ax.text(
                (x1 + x2) / 2, (y1 + y2) / 2 + 0.018,
                text,
                ha="center", va="bottom",
                fontsize=9, color="#555555"
            )

    # --------------------------------------------------
    # Main step boxes
    # --------------------------------------------------
    box_w = 0.15
    box_h = 0.17
    y_box = 0.55
    xs = [0.02, 0.215, 0.41, 0.605, 0.80]

    add_box(
        xs[0], y_box, box_w, box_h,
        "STEP 1\n\nFilter candidate actors\nby stage, source and\nindustrial criteria\n(EPRTR / UWWTPS,\nNACE codes, capacity)",
        fc="#EAF7EC", ec=theme_color, fontsize=10, weight="bold"
    )

    add_box(
        xs[1], y_box, box_w, box_h,
        "STEP 2\n\nChoose a pivot stage\nand build one BallTree\nspatial index per stage\nfrom actor coordinates",
        fc="#EAF7EC", ec=theme_color, fontsize=10, weight="bold"
    )

    add_box(
        xs[2], y_box, box_w, box_h,
        "STEP 3\n\nSearch neighbours\nwithin the maximum\nallowed distance\nbetween consecutive\nstages",
        fc="#EAF7EC", ec=theme_color, fontsize=10, weight="bold"
    )

    add_box(
        xs[3], y_box, box_w, box_h,
        "STEP 4\n\nCheck model rules:\nminimum actors,\nAND / OR logic,\nsource filters and\ndistance constraints",
        fc="#EAF7EC", ec=theme_color, fontsize=10, weight="bold"
    )

    add_box(
        xs[4], y_box, box_w, box_h,
        "STEP 5\n\nValidate one potential\nreplication network,\ncompute its centroid,\nand export results\nto map / Excel / PDF",
        fc="#EAF7EC", ec=theme_color, fontsize=10, weight="bold"
    )

    for i in range(4):
        add_arrow(xs[i] + box_w, y_box + box_h / 2, xs[i + 1], y_box + box_h / 2)

    # --------------------------------------------------
    # Small illustrations under each step
    # --------------------------------------------------

    # Step 1: actor filtering
    x0 = xs[0]
    ax.text(x0 + box_w / 2, 0.45, "Example filtered stages", ha="center", va="center",
            fontsize=8, color="#555555", weight="bold")
    add_box(x0 + 0.01, 0.34, 0.04, 0.07, "10.41", fc="#FDE9E7", fontsize=8)
    add_box(x0 + 0.06, 0.34, 0.04, 0.07, "10.42", fc="#FDE9E7", fontsize=8)
    add_box(x0 + 0.11, 0.34, 0.04, 0.07, "UWW", fc="#FFF2CC", fontsize=8)
    add_box(x0 + 0.06, 0.24, 0.04, 0.07, "20.16", fc="#DDEBF7", fontsize=8)

    # Step 2: BallTree concept
    x0 = xs[1]
    ax.text(x0 + box_w / 2, 0.45, "Spatial indexing with BallTree", ha="center", va="center",
            fontsize=8, color="#555555", weight="bold")
    pts = [(x0 + 0.04, 0.29), (x0 + 0.07, 0.37), (x0 + 0.10, 0.31), (x0 + 0.13, 0.36)]
    for px, py in pts:
        ax.add_patch(Circle((px, py), 0.006, facecolor=theme_color, edgecolor="black", linewidth=0.4))
    ax.plot([pts[0][0], pts[1][0]], [pts[0][1], pts[1][1]], color="#888888", lw=1)
    ax.plot([pts[1][0], pts[2][0]], [pts[1][1], pts[2][1]], color="#888888", lw=1)
    ax.plot([pts[1][0], pts[3][0]], [pts[1][1], pts[3][1]], color="#888888", lw=1)

    # Step 3: radius search
    x0 = xs[2]
    ax.text(x0 + box_w / 2, 0.45, "Radius search around a pivot actor", ha="center", va="center",
            fontsize=8, color="#555555", weight="bold")
    pivot = (x0 + 0.08, 0.32)
    ax.add_patch(Circle(pivot, 0.007, facecolor="#FF6F61", edgecolor="black", linewidth=0.5))
    ax.add_patch(Circle(pivot, 0.05, facecolor="none", edgecolor="#888888", linewidth=1.2, linestyle="--"))
    neigh = [(x0 + 0.05, 0.35), (x0 + 0.11, 0.35), (x0 + 0.12, 0.29), (x0 + 0.04, 0.28)]
    for px, py in neigh:
        ax.add_patch(Circle((px, py), 0.006, facecolor=theme_color, edgecolor="black", linewidth=0.4))
    ax.text(x0 + 0.08, 0.245, "query_radius(max_distance)", ha="center", va="center", fontsize=8, color="#555555")

    # Step 4: rules
    x0 = xs[3]
    ax.text(x0 + box_w / 2, 0.45, "Rule validation", ha="center", va="center",
            fontsize=8, color="#555555", weight="bold")
    add_box(x0 + 0.02, 0.35, 0.12, 0.045, "AND / OR logic", fc="#F3F3F3", fontsize=8)
    add_box(x0 + 0.02, 0.29, 0.12, 0.045, "Min actors", fc="#F3F3F3", fontsize=8)
    add_box(x0 + 0.02, 0.23, 0.12, 0.045, "Max distance", fc="#F3F3F3", fontsize=8)

    # Step 5: final network
    x0 = xs[4]
    ax.text(x0 + box_w / 2, 0.45, "Validated multi-stage network", ha="center", va="center",
            fontsize=8, color="#555555", weight="bold")
    p1 = (x0 + 0.03, 0.34)
    p2 = (x0 + 0.08, 0.30)
    p3 = (x0 + 0.13, 0.36)
    p4 = (x0 + 0.13, 0.25)
    for px, py, c in [(p1[0], p1[1], "#D9534F"), (p2[0], p2[1], "#F0AD4E"), (p3[0], p3[1], "#5BC0DE"), (p4[0], p4[1], "#5BC0DE")]:
        ax.add_patch(Circle((px, py), 0.007, facecolor=c, edgecolor="black", linewidth=0.5))
    ax.annotate("", xy=p2, xytext=p1, arrowprops=dict(arrowstyle="->", lw=1.5, color="#444444"))
    ax.annotate("", xy=p3, xytext=p2, arrowprops=dict(arrowstyle="->", lw=1.5, color="#444444"))
    ax.annotate("", xy=p4, xytext=p2, arrowprops=dict(arrowstyle="->", lw=1.5, color="#444444"))
    ax.add_patch(Circle((x0 + 0.10, 0.31), 0.01, facecolor=theme_color, edgecolor="black", linewidth=0.6))
    ax.text(x0 + 0.10, 0.285, "centroid", ha="center", va="top", fontsize=8, color="#555555")

    # --------------------------------------------------
    # Explanatory caption
    # --------------------------------------------------
    caption = (
        "Principle of the replication detection algorithm. Candidate actors are first filtered by stage and "
        "industrial criteria. A pivot stage is then selected, and BallTree spatial indexing is used to efficiently "
        "identify nearby actors within the maximum allowed distance between stages. Multi-stage configurations are "
        "validated against actor-group, minimum-count, and AND/OR logic constraints to detect potential replication networks."
    )

    fig.text(
        0.05, 0.06,
        caption,
        fontsize=10,
        color="#555555",
        ha="left",
        va="bottom",
        wrap=True
    )

    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)

# ======================================================
# EXPORT TABLES
# ======================================================
def build_export_tables(paths, layer_dfs, layer_defs):
    network_rows = []
    actor_rows = []

    for p in paths:
        row = {
            "network_id": int(p.get("path_id", -1)),
            "pivot_stage": int(p.get("pivot_layer", -1)) + 1,
            "pivot_idx": int(p.get("pivot_idx", -1)),
            "centroid_lat": float(p.get("lat", np.nan)),
            "centroid_lon": float(p.get("lon", np.nan)),
        }

        total_actors = 0
        countries = set()

        piv_layer = int(p.get("pivot_layer", -1))
        piv_idx = int(p.get("pivot_idx", -1))

        if (
            piv_layer >= 0
            and layer_dfs is not None
            and piv_layer < len(layer_dfs)
            and piv_idx < len(layer_dfs[piv_layer])
        ):
            piv = layer_dfs[piv_layer].iloc[piv_idx]
            row["pivot_name"] = piv.get("__name", "")
            row["pivot_code"] = piv.get("__layer_code", "")
            row["pivot_source"] = piv.get("__layer_source", "")
            row["pivot_country"] = get_actor_country(piv)
        else:
            row["pivot_name"] = ""
            row["pivot_code"] = ""
            row["pivot_source"] = ""
            row["pivot_country"] = ""

        signature = []

        for li, idxs in enumerate(p["layer_sets"]):
            n = 0 if idxs is None else len(idxs)
            row[f"n_stage_{li+1}"] = n
            total_actors += n
            signature.append(f"S{li+1}:{n}")

            if idxs is None:
                continue

            dfL = layer_dfs[li]
            for ridx in idxs:
                actor = dfL.iloc[int(ridx)]
                country = get_actor_country(actor)
                countries.add(country)

                actor_rows.append({
                    "network_id": int(p.get("path_id", -1)),
                    "stage": li + 1,
                    "stage_label": layer_defs[li]["label"],
                    "actor_role": "pivot" if (li == piv_layer and int(ridx) == piv_idx) else "member",
                    "source": actor.get("__layer_source", ""),
                    "code": actor.get("__layer_code", ""),
                    "name": actor.get("__name", ""),
                    "country": country,
                    "latitude": float(actor.latitude),
                    "longitude": float(actor.longitude),
                    "is_pivot": (li == piv_layer and int(ridx) == piv_idx),
                })

        row["total_actors"] = total_actors
        row["network_countries"] = ", ".join(sorted(c for c in countries if c and c != "NAN"))
        row["n_countries"] = len([c for c in countries if c and c != "NAN"])
        row["stage_signature"] = " | ".join(signature)
        network_rows.append(row)

    df_networks = pd.DataFrame(network_rows)
    df_actors = pd.DataFrame(actor_rows)
    return df_networks, df_actors


def build_links_table(paths, layer_dfs, layer_defs):
    rows = []

    for p in paths:
        network_id = int(p["path_id"])
        layer_sets = p["layer_sets"]

        for li in range(len(layer_sets) - 1):
            idxs_src = layer_sets[li]
            idxs_tgt = layer_sets[li + 1]

            if not idxs_src or not idxs_tgt:
                continue

            df_src = layer_dfs[li]
            df_tgt = layer_dfs[li + 1].iloc[idxs_tgt].copy().reset_index(drop=True)

            tree_tgt = BallTree(
                np.radians(df_tgt[["latitude", "longitude"]].values),
                metric="haversine"
            )

            # Requête vectorisée : tous les acteurs source de l'étape sont
            # interrogés en une seule fois (au lieu d'une requête BallTree
            # par acteur dans une boucle Python). Résultat strictement
            # identique, beaucoup plus rapide sur les grands réseaux.
            df_src_sel = df_src.iloc[idxs_src]
            src_coords = np.radians(df_src_sel[["latitude", "longitude"]].values)
            dist_all, ind_all = tree_tgt.query(src_coords, k=1)

            for row_i, ridx in enumerate(idxs_src):
                src_row = df_src_sel.iloc[row_i]
                src_lat = float(src_row.latitude)
                src_lon = float(src_row.longitude)

                tgt_row = df_tgt.iloc[int(ind_all[row_i][0])]
                dist_km = float(dist_all[row_i][0]) * KMS_PER_RADIAN

                rows.append({
                    "network_id": network_id,
                    "from_stage": li + 1,
                    "to_stage": li + 2,
                    "from_stage_label": layer_defs[li]["label"],
                    "to_stage_label": layer_defs[li + 1]["label"],
                    "from_name": src_row.get("__name", ""),
                    "to_name": tgt_row.get("__name", ""),
                    "from_code": src_row.get("__layer_code", ""),
                    "to_code": tgt_row.get("__layer_code", ""),
                    "from_source": src_row.get("__layer_source", ""),
                    "to_source": tgt_row.get("__layer_source", ""),
                    "from_country": get_actor_country(src_row),
                    "to_country": get_actor_country(tgt_row),
                    "distance_km": round(dist_km, 2),
                    "link_type": "nearest"
                })

    return pd.DataFrame(rows)


def build_parameters_table(layer_defs, pivot_layer, max_neighbors, max_paths):
    rows = []
    rows.append({"parameter": "pivot_layer", "value": pivot_layer + 1})
    rows.append({"parameter": "max_neighbors", "value": max_neighbors})
    rows.append({"parameter": "max_paths", "value": max_paths})

    for i, ld in enumerate(layer_defs):
        rows.append({"parameter": f"stage_{i+1}_label", "value": ld.get("label", "")})
        rows.append({"parameter": f"stage_{i+1}_source", "value": ld.get("source", "")})
        rows.append({"parameter": f"stage_{i+1}_logic", "value": ld.get("logic", "")})
        rows.append({"parameter": f"stage_{i+1}_max_to_next_km", "value": ld.get("max_to_next_km", "")})
        if ld.get("source") == "UWWTPS":
            rows.append({"parameter": f"stage_{i+1}_cap_min", "value": ld.get("cap_min", "")})
            rows.append({"parameter": f"stage_{i+1}_cap_max", "value": ld.get("cap_max", "")})

        for gi, g in enumerate(ld.get("actor_groups", [])):
            rows.append({"parameter": f"stage_{i+1}_group_{gi+1}_label", "value": g.get("label", "")})
            rows.append({"parameter": f"stage_{i+1}_group_{gi+1}_codes", "value": ", ".join(g.get("codes", []))})
            rows.append({"parameter": f"stage_{i+1}_group_{gi+1}_min_actors", "value": g.get("min_actors", 1)})

    return pd.DataFrame(rows)


def prepare_export_artifacts(paths, layer_dfs, layer_defs, counts, pivot_layer, max_neighbors, max_paths):
    networks_df, actors_df = build_export_tables(paths, layer_dfs, layer_defs)
    links_df = build_links_table(paths, layer_dfs, layer_defs)

    if len(counts) > 0:
        country_stats_df = counts.reset_index()
        country_stats_df.columns = ["country", "n_networks"]
        country_stats_df["share_networks_pct"] = (
            country_stats_df["n_networks"] / max(len(networks_df), 1) * 100
        ).round(2)

        if not actors_df.empty:
            distinct_actors_by_country = (
                actors_df.groupby("country")["name"]
                .nunique()
                .reset_index(name="n_distinct_actors")
            )
            pivot_networks = (
                networks_df.groupby("pivot_country")["network_id"]
                .count()
                .reset_index(name="n_pivot_networks")
                .rename(columns={"pivot_country": "country"})
            )
            country_stats_df = country_stats_df.merge(distinct_actors_by_country, on="country", how="left")
            country_stats_df = country_stats_df.merge(pivot_networks, on="country", how="left")
        else:
            country_stats_df["n_distinct_actors"] = 0
            country_stats_df["n_pivot_networks"] = 0

        country_stats_df["n_distinct_actors"] = country_stats_df["n_distinct_actors"].fillna(0).astype(int)
        country_stats_df["n_pivot_networks"] = country_stats_df["n_pivot_networks"].fillna(0).astype(int)
    else:
        country_stats_df = pd.DataFrame(
            columns=["country", "n_networks", "share_networks_pct", "n_distinct_actors", "n_pivot_networks"]
        )

    summary_df = pd.DataFrame({
        "indicator": [
            "Generated at",
            "Number of networks",
            "Number of actors",
            "Number of countries represented",
            "Average actors per network",
            "Median actors per network",
            "Average links per network",
        ],
        "value": [
            datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
            len(networks_df),
            len(actors_df),
            0 if country_stats_df.empty else len(country_stats_df),
            round(networks_df["total_actors"].mean(), 2) if not networks_df.empty else 0,
            round(networks_df["total_actors"].median(), 2) if not networks_df.empty else 0,
            round(len(links_df) / max(len(networks_df), 1), 2) if not networks_df.empty else 0,
        ]
    })

    parameters_df = build_parameters_table(layer_defs, pivot_layer, max_neighbors, max_paths)

    readme_df = pd.DataFrame({
        "field": [
            "Purpose",
            "Definition of a detected network",
            "Pivot logic",
            "Networks sheet",
            "Actors sheet",
            "Links sheet",
            "Country_stats sheet",
            "Parameters sheet",
        ],
        "description": [
            "Replication analysis package for circular economy value chain screening.",
            "A detected network is a multi-stage geospatial configuration matching selected actor groups and distance constraints.",
            "Any stage may be used as pivot; network construction is propagated bidirectionally from that pivot.",
            "One row per detected network.",
            "One row per actor included in a detected network.",
            "One row per nearest actor-to-actor link between consecutive stages.",
            "Aggregated distribution of detected networks by country.",
            "Main model parameters used for the run."
        ]
    })

    return ExportArtifacts(
        summary_df=summary_df,
        networks_df=networks_df,
        actors_df=actors_df,
        links_df=links_df,
        country_stats_df=country_stats_df,
        parameters_df=parameters_df,
        readme_df=readme_df,
    )


# ======================================================
# EXPORT FILE BUILDERS
# ======================================================
def build_overview_map_html(paths, layer_dfs, layer_defs, centroids_df, heatmap_data=None, show_heatmap=False):
    import json
    import math
    from folium.plugins import HeatMap

    if centroids_df is None or len(centroids_df) == 0:
        m = folium.Map(location=[50, 10], zoom_start=4, tiles="CartoDB positron")
        return m.get_root().render().encode("utf-8")

    lat0 = float(centroids_df["lat"].mean())
    lon0 = float(centroids_df["lon"].mean())

    m = make_base_map(lat0, lon0, zoom=6)
    map_name = m.get_name()

    # --------------------------------------------------
    # Layer groups
    # --------------------------------------------------
    centroid_group = folium.FeatureGroup(name="Network centroids", show=True)
    centroid_group.add_to(m)

    if heatmap_data:
        heatmap_group = folium.FeatureGroup(name="Hotspot heatmap", show=show_heatmap)
        HeatMap(
            heatmap_data,
            radius=35,
            blur=25,
            min_opacity=0.6,
            max_zoom=8
        ).add_to(heatmap_group)
        heatmap_group.add_to(m)

    # --------------------------------------------------
    # Helpers
    # --------------------------------------------------
    def actor_type_label(li, src, n_layers):
        if src == "UWWTPS":
            return "Intermediate"
        if li == 0:
            return "Emitter"
        if li == n_layers - 1:
            return "User"
        return "Intermediate"

    def color_for_layer(li, src, n_layers):
        if src == "UWWTPS":
            return "orange"
        if li == 0:
            return "red"
        if li == n_layers - 1:
            return "blue"
        return "purple"

    # --------------------------------------------------
    # Prepare network payload for JS
    # --------------------------------------------------
    network_payload = {}

    for p in paths:
        pid = int(p["path_id"])
        layer_sets = p["layer_sets"]

        actors = []
        links = []
        bounds = []

        # Actors
        for li, idxs in enumerate(layer_sets):
            if idxs is None:
                continue

            dfL = layer_dfs[li]
            for ridx in idxs:
                row = dfL.iloc[int(ridx)]
                lat = float(row.latitude)
                lon = float(row.longitude)
                src = row.get("__layer_source", "")
                actor_type = actor_type_label(li, src, len(layer_sets))
                color = color_for_layer(li, src, len(layer_sets))

                bounds.append([lat, lon])

                popup_html = (
                    f"<b>{actor_type}</b><br>"
                    f"Stage: {li+1} — {layer_defs[li]['label']}<br>"
                    f"Name: {row.get('__name', '')}<br>"
                    f"Code: {row.get('__layer_code', '')}<br>"
                    f"Source: {row.get('__layer_source', '')}"
                )

                actors.append({
                    "lat": lat,
                    "lon": lon,
                    "color": color,
                    "radius": 5,
                    "tooltip": f"{actor_type} • {row.get('__layer_code', '')}",
                    "popup": popup_html,
                })

        # Links: actor -> nearest actor in next stage
        for li in range(len(layer_sets) - 1):
            idxs_src = layer_sets[li]
            idxs_tgt = layer_sets[li + 1]

            if not idxs_src or not idxs_tgt:
                continue

            df_src = layer_dfs[li]
            df_tgt_full = layer_dfs[li + 1]
            df_tgt = df_tgt_full.iloc[idxs_tgt].copy().reset_index(drop=True)

            if len(df_tgt) == 0:
                continue

            tree_tgt = BallTree(
                np.radians(df_tgt[["latitude", "longitude"]].values),
                metric="haversine"
            )

            # Requête vectorisée sur tous les acteurs source de l'étape
            # (cf. build_links_table) : même résultat, sans boucle Python
            # par acteur pour interroger le BallTree.
            df_src_sel = df_src.iloc[idxs_src]
            src_coords = np.radians(df_src_sel[["latitude", "longitude"]].values)
            dist_all, ind_all = tree_tgt.query(src_coords, k=1)

            for row_i, ridx in enumerate(idxs_src):
                src_row = df_src_sel.iloc[row_i]
                src_lat = float(src_row.latitude)
                src_lon = float(src_row.longitude)

                tgt_row = df_tgt.iloc[int(ind_all[row_i][0])]

                tgt_lat = float(tgt_row.latitude)
                tgt_lon = float(tgt_row.longitude)

                links.append({
                    "coords": [[src_lat, src_lon], [tgt_lat, tgt_lon]],
                    "color": "black",
                    "weight": 1.5,
                    "opacity": 0.6
                })

        network_payload[str(pid)] = {
            "centroid": [float(p["lat"]), float(p["lon"])],
            "bounds": bounds,
            "actors": actors,
            "links": links,
        }

    payload_json = json.dumps(network_payload)

    # --------------------------------------------------
    # Centroid markers
    # --------------------------------------------------
    for _, r in centroids_df.iterrows():
        pid = int(r["path_id"])

        popup_html = f"""
        <div style="min-width:180px">
            <b>Network {pid}</b><br>
            <button onclick="showNetwork('{pid}')" style="margin-top:6px; cursor:pointer;">
                Show network
            </button>
            <button onclick="hideNetwork()" style="margin-top:6px; margin-left:6px; cursor:pointer;">
                Hide network
            </button>
        </div>
        """

        folium.CircleMarker(
            location=[float(r["lat"]), float(r["lon"])],
            radius=5,
            color="#00B400",
            fill=True,
            fill_opacity=0.9,
            tooltip=f"Network {pid}",
            popup=folium.Popup(popup_html, max_width=260)
        ).add_to(centroid_group)

    # --------------------------------------------------
    # Layer control
    # --------------------------------------------------
    folium.LayerControl(collapsed=False).add_to(m)

    # --------------------------------------------------
    # Legend
    # --------------------------------------------------
    legend_html = """
    <div style="
        position: fixed;
        bottom: 40px;
        left: 40px;
        z-index: 9999;
        background-color: white;
        border: 2px solid #666;
        border-radius: 6px;
        padding: 10px 12px;
        font-size: 13px;
        box-shadow: 2px 2px 6px rgba(0,0,0,0.2);
    ">
        <div style="font-weight: bold; margin-bottom: 8px;">Legend</div>

        <div style="margin-bottom: 4px;">
            <span style="
                display:inline-block;
                width:10px;
                height:10px;
                border-radius:50%;
                background:green;
                margin-right:8px;"></span>
            Potential replicated pilots (click on it to display the symbiosis network)
        </div>

        <div style="margin-bottom: 4px;">
            <span style="
                display:inline-block;
                width:10px;
                height:10px;
                border-radius:50%;
                background:red;
                margin-right:8px;"></span>
            Emitter
        </div>

        <div style="margin-bottom: 4px;">
            <span style="
                display:inline-block;
                width:10px;
                height:10px;
                border-radius:50%;
                background:orange;
                margin-right:8px;"></span>
            Intermediate
        </div>

        <div style="margin-bottom: 4px;">
            <span style="
                display:inline-block;
                width:10px;
                height:10px;
                border-radius:50%;
                background:blue;
                margin-right:8px;"></span>
            End-user
        </div>

        <div style="margin-top: 8px;">
            <span style="
                display:inline-block;
                width:18px;
                height:2px;
                background:black;
                margin-right:8px;
                vertical-align:middle;"></span>
            Link bewteen two industrial sites
        </div>
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))

    # --------------------------------------------------
    # Offline CSS for arrow heads
    # --------------------------------------------------
    arrow_css = """
    <style>
    .flow-arrow-icon {
        width: 0;
        height: 0;
        border-left: 8px solid transparent;
        border-right: 8px solid transparent;
        border-bottom: 14px solid black;
        transform-origin: center center;
    }
    </style>
    """
    m.get_root().header.add_child(folium.Element(arrow_css))

    # --------------------------------------------------
    # JS: display/hide network + arrows without plugin
    # --------------------------------------------------
    custom_js = f"""
    <script>
    var networkData = {payload_json};
    var activeNetworkLayer = null;

    function computeBearing(lat1, lon1, lat2, lon2) {{
        var rad = Math.PI / 180.0;
        var phi1 = lat1 * rad;
        var phi2 = lat2 * rad;
        var dLon = (lon2 - lon1) * rad;

        var y = Math.sin(dLon) * Math.cos(phi2);
        var x = Math.cos(phi1) * Math.sin(phi2) -
                Math.sin(phi1) * Math.cos(phi2) * Math.cos(dLon);

        var brng = Math.atan2(y, x) * 180.0 / Math.PI;
        return (brng + 360.0) % 360.0;
    }}

    function addArrowMarker(group, coords) {{
        if (!coords || coords.length < 2) return;

        var p1 = coords[0];
        var p2 = coords[1];

        var midLat = (p1[0] + p2[0]) / 2.0;
        var midLon = (p1[1] + p2[1]) / 2.0;

        var bearing = computeBearing(p1[0], p1[1], p2[0], p2[1]);

        var html = '<div class="flow-arrow-icon" style="transform: rotate(' + bearing + 'deg);"></div>';

        var icon = L.divIcon({{
            className: '',
            html: html,
            iconSize: [16, 16],
            iconAnchor: [8, 8]
        }});

        L.marker([midLat, midLon], {{
            icon: icon,
            interactive: false
        }}).addTo(group);
    }}

    function buildNetworkLayer(pid) {{
        var d = networkData[String(pid)];
        if (!d) return null;

        var group = L.layerGroup();

        d.actors.forEach(function(a) {{
            var marker = L.circleMarker([a.lat, a.lon], {{
                radius: a.radius || 5,
                color: a.color || "blue",
                fillColor: a.color || "blue",
                fillOpacity: 0.85,
                opacity: 1,
                weight: 1
            }});

            if (a.tooltip) marker.bindTooltip(a.tooltip);
            if (a.popup) marker.bindPopup(a.popup);

            marker.addTo(group);
        }});

        d.links.forEach(function(l) {{
            var line = L.polyline(l.coords, {{
                color: l.color || "black",
                weight: l.weight || 1.5,
                opacity: l.opacity || 0.6
            }});
            line.addTo(group);

            addArrowMarker(group, l.coords);
        }});

        return group;
    }}

    function hideNetwork() {{
        var map = {map_name};
        if (activeNetworkLayer) {{
            map.removeLayer(activeNetworkLayer);
            activeNetworkLayer = null;
        }}
    }}

    function showNetwork(pid) {{
        var map = {map_name};

        hideNetwork();

        var layer = buildNetworkLayer(pid);
        if (!layer) return;

        activeNetworkLayer = layer;
        activeNetworkLayer.addTo(map);

        var d = networkData[String(pid)];
        if (d && d.bounds && d.bounds.length > 0) {{
            map.fitBounds(d.bounds, {{padding: [30, 30]}});
        }}
    }}

    window.showNetwork = showNetwork;
    window.hideNetwork = hideNetwork;
    </script>
    """
    m.get_root().html.add_child(folium.Element(custom_js))

    return m.get_root().render().encode("utf-8")


def add_excel_table(ws, df, table_name):
    if df.empty:
        return
    last_row = ws.max_row
    last_col = ws.max_column
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def format_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len("" if c.value is None else str(c.value)) for c in col)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def build_excel_bytes(artifacts: ExportArtifacts):
    bio = io.BytesIO()

    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        artifacts.readme_df.to_excel(writer, sheet_name="README", index=False)
        artifacts.summary_df.to_excel(writer, sheet_name="Summary", index=False)
        artifacts.networks_df.to_excel(writer, sheet_name="Networks", index=False)
        artifacts.actors_df.to_excel(writer, sheet_name="Actors", index=False)
        artifacts.links_df.to_excel(writer, sheet_name="Links", index=False)
        artifacts.country_stats_df.to_excel(writer, sheet_name="Country_stats", index=False)
        artifacts.parameters_df.to_excel(writer, sheet_name="Parameters", index=False)

        wb = writer.book

        for ws in wb.worksheets:
            format_worksheet(ws)

        add_excel_table(wb["README"], artifacts.readme_df, "TblReadme")
        add_excel_table(wb["Summary"], artifacts.summary_df, "TblSummary")
        add_excel_table(wb["Networks"], artifacts.networks_df, "TblNetworks")
        add_excel_table(wb["Actors"], artifacts.actors_df, "TblActors")
        add_excel_table(wb["Links"], artifacts.links_df, "TblLinks")
        add_excel_table(wb["Country_stats"], artifacts.country_stats_df, "TblCountryStats")
        add_excel_table(wb["Parameters"], artifacts.parameters_df, "TblParameters")

    bio.seek(0)
    return bio.getvalue()


def draw_report_header(fig, title, subtitle):
    theme_color = "#02A717"

    # --------------------------------------------------
    # TITLE
    # --------------------------------------------------
    fig.text(
        0.05,
        0.965,
        title,
        fontsize=26,
        weight="bold",
        color=theme_color,
        va="top"
    )

    # --------------------------------------------------
    # SUBTITLE
    # --------------------------------------------------
    fig.text(
        0.05,
        0.925,
        subtitle,
        fontsize=14,
        color="#555555",
        va="top"
    )

    # --------------------------------------------------
    # GREEN LINE
    # --------------------------------------------------
    fig.lines.append(
        plt.Line2D(
            [0.05, 0.95],
            [0.895, 0.895],
            transform=fig.transFigure,
            color=theme_color,
            linewidth=3
        )
    )

    # --------------------------------------------------
    # LOGO
    # --------------------------------------------------
    try:
        logo = mpimg.imread("data/Strane-logo.png")

        ax_logo = fig.add_axes([0.82, 0.905, 0.12, 0.07])
        ax_logo.imshow(logo)
        ax_logo.axis("off")

    except Exception:
        pass


def build_pdf_bytes(centroids_df, counts, artifacts: ExportArtifacts, paths, layer_dfs, layer_defs):
    bio = io.BytesIO()

    with PdfPages(bio) as pdf:
        # Page 1 - Cover
        fig = plt.figure(figsize=(11.69, 8.27))
        plt.axis("off")
        draw_report_header(fig, "Replication Analysis Package", "Circular economy value chains across Europe")
        fig.text(0.05, 0.78, "Synthesis report", fontsize=28, weight="bold", color="#1F4E78")
        fig.text(0.05, 0.68, "Partner-ready overview of detected industrial symbiosis networks", fontsize=14)
        fig.text(0.05, 0.58, f"Detected networks: {len(artifacts.networks_df)}", fontsize=14)
        fig.text(0.05, 0.53, f"Actors involved: {len(artifacts.actors_df)}", fontsize=14)
        fig.text(0.05, 0.48, f"Countries represented: {len(artifacts.country_stats_df)}", fontsize=14)
        fig.text(0.05, 0.12, "Generated automatically from the replication screening tool.", fontsize=10, color="#777777")
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Page 2 - Executive summary
        fig = plt.figure(figsize=(11.69, 8.27))
        plt.axis("off")
        draw_report_header(fig, "Executive summary", "Key indicators")
        y = 0.78
        for _, row in artifacts.summary_df.iterrows():
            fig.text(0.08, y, f"{row['indicator']} :", fontsize=12, weight="bold", color="#1F4E78")
            fig.text(0.38, y, f"{row['value']}", fontsize=12)
            y -= 0.08

        fig.text(
            0.08, 0.18,
            "This report summarises the replication analysis of the selected circular value chain, "
            "including the spatial distribution of detected network centroids, country distribution, "
            "and example identified networks.",
            fontsize=11, wrap=True
        )
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Page 2 - Methodological note
        draw_balltree_method_page(pdf)

        # Page 3 - Map of centroids on OSM basemap
        draw_centroid_map_page(pdf, centroids_df)

        # Page 4 - Heatmap of centroids on OSM basemap
        draw_centroid_heatmap_page(pdf, centroids_df)

        # Page 5 - Country histogram
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        draw_report_header(fig, "Country distribution", "Distribution of potential replicated pilots per country")
        if len(counts) > 0:
            counts.plot(kind="bar", ax=ax, color="#02A717", edgecolor="black")
            ax.set_xlabel("Country")
            ax.set_ylabel("Number of potential replicated pilots")
            ax.grid(axis="y", alpha=0.3)
        else:
            ax.text(0.5, 0.5, "No country data available", ha="center", va="center")
            ax.axis("off")
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Page 7 - Methodology flow diagram
        draw_methodology_flow_page(pdf, layer_defs)

        # Page 8 - Example detected network diagram
        draw_example_network_page(
            pdf,
            paths=paths,
            layer_dfs=layer_dfs,
            layer_defs=layer_defs,
            network_id=0
        )

    bio.seek(0)
    return bio.getvalue()


def build_manifest_text(artifacts: ExportArtifacts):
    return (
        "Replication analysis package\n"
        f"Generated at: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n"
        f"Networks: {len(artifacts.networks_df)}\n"
        f"Actors: {len(artifacts.actors_df)}\n"
        f"Links: {len(artifacts.links_df)}\n"
        f"Countries: {len(artifacts.country_stats_df)}\n"
    )


def build_zip_bytes(html_bytes, excel_bytes, pdf_bytes, artifacts: ExportArtifacts):
    bio = io.BytesIO()
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    prefix = f"replication_analysis_{stamp}"

    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{prefix}_map.html", html_bytes)
        zf.writestr(f"{prefix}_networks.xlsx", excel_bytes)
        zf.writestr(f"{prefix}_summary.pdf", pdf_bytes)
        zf.writestr(f"{prefix}_manifest.txt", build_manifest_text(artifacts))

    bio.seek(0)
    return bio.getvalue()


# ======================================================
# MIGRATION / SECURISATION layer_defs (V2)
# ======================================================
for ld in st.session_state.layer_defs:
    if "logic" not in ld:
        ld["logic"] = "AND"

    if "actor_groups" not in ld:
        if ld.get("source") == "EPRTR":
            codes = []
            if "codes" in ld and ld["codes"]:
                codes = [c.strip() for c in str(ld["codes"]).split(",")]
            ld["actor_groups"] = [{
                "label": "Acteurs",
                "codes": codes,
                "min_actors": int(ld.get("min_actors", 1)),
            }]
        else:
            ld["actor_groups"] = [{
                "label": "Acteurs",
                "codes": ["UWWTPS"],
                "min_actors": int(ld.get("min_actors", 1)),
            }]

    ld.pop("codes", None)
    ld.pop("min_actors", None)


# ======================================================
# SIDEBAR
# ======================================================
with st.sidebar:
    st.header("Data")
    eprtr_path = st.text_input("EPRTR.csv", "EPRTR.csv")
    uww_path = st.text_input("UWWTPS.csv", "UWWTPS.csv")

    st.divider()
    st.header("Display")

    show_heatmap = st.checkbox(
        "Afficher la heatmap des réseaux",
        value=False,
        help="Heatmap basée sur les centroïdes des réseaux détectés"
    )

    max_neighbors = st.slider("Max acteurs conservés par étage", 5, 500, 200, 5)
    max_paths = st.slider("Max réseaux conservés", 100, 50000, 20000, 100)

    st.divider()
    st.header("Pivot")
    pivot_layer = st.selectbox(
        "Pivot Stage",
        options=list(range(len(st.session_state.layer_defs))),
        index=min(max(st.session_state.pivot_layer, 0), len(st.session_state.layer_defs) - 1),
        format_func=lambda i: f"Stage {i+1} — {st.session_state.layer_defs[i].get('label','')}"
    )
    st.session_state.pivot_layer = int(pivot_layer)

    pivot_mode = st.selectbox(
        "Calculation method",
        options=["Pivot stakeholder (selection)", "All stakeholder in the pivot"],
        index=0 if st.session_state.pivot_mode == "Pivot stakeholder (selection)" else 1
    )
    st.session_state.pivot_mode = pivot_mode

    pivot_search = st.text_input(
        "Filter the pivots list (text)",
        value="",
        help="Filter by name / code / country to quickly find a pivot."
    )

    st.divider()
    st.header("Stages")
    st.caption("The maximum distance is applied between adjacent stages (i → i+1).")
    st.caption("min_actors = the minimum number of actors required within the radius around the centroid of the neighbouring stage.")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Add a stage"):
            st.session_state.layer_defs.append({
                "label": f"Stage {len(st.session_state.layer_defs)+1}",
                "source": "EPRTR",
                "codes": "",
                "min_actors": 1,
                "max_to_next_km": 30.0,
            })
            st.rerun()
    with c2:
        if st.button("Delete last stage"):
            if len(st.session_state.layer_defs) > 2:
                st.session_state.layer_defs.pop()
                st.session_state.pivot_layer = min(
                    st.session_state.pivot_layer,
                    len(st.session_state.layer_defs) - 1
                )
                st.rerun()

    for i, ld in enumerate(st.session_state.layer_defs):
        st.markdown(f"### Stage {i+1}")
        ld["label"] = st.text_input(
            f"Label (stage {i+1})",
            ld.get("label", f"stage {i+1}"),
            key=f"lbl_{i}"
        )
        ld["source"] = st.selectbox(
            f"Source (stage {i+1})",
            ["EPRTR", "UWWTPS"],
            index=0 if ld.get("source", "EPRTR") == "EPRTR" else 1,
            key=f"src_{i}"
        )

        if ld["source"] == "UWWTPS":
            c1, c2 = st.columns(2)

            with c1:
                ld["cap_min"] = st.number_input(
                    "Minimal capacity UWWTPS",
                    min_value=0.0,
                    value=float(ld.get("cap_min", 0.0)),
                    step=1000.0,
                    key=f"cap_min_{i}"
                )

            with c2:
                ld["cap_max"] = st.number_input(
                    "Maximal capacity UWWTPS",
                    min_value=float(ld.get("cap_min", 0.0)),
                    value=float(ld.get("cap_max", 30000.0)),
                    step=1000.0,
                    key=f"cap_max_{i}"
                )

        ld["logic"] = st.selectbox(
            "Logic between types of stakeholders",
            ["AND", "OR"],
            index=0 if ld.get("logic", "AND") == "AND" else 1,
            key=f"logic_{i}"
        )

        st.markdown("**Types of stakeholders in this stage**")

        for gi, g in enumerate(ld["actor_groups"]):
            c1, c2, c3 = st.columns([2, 3, 1])

            with c1:
                g["label"] = st.text_input(
                    "Nom",
                    g.get("label", ""),
                    key=f"g_label_{i}_{gi}"
                )

            with c2:
                g["codes"] = st.text_input(
                    "Codes (separated by ,)",
                    ",".join(g.get("codes", [])),
                    key=f"g_codes_{i}_{gi}"
                ).replace(" ", "").split(",")

            with c3:
                g["min_actors"] = st.number_input(
                    "Min",
                    min_value=1,
                    value=int(g.get("min_actors", 1)),
                    step=1,
                    key=f"g_min_{i}_{gi}"
                )

        if st.button("➕ Add another type of stakehodlers", key=f"add_group_{i}"):
            ld["actor_groups"].append({
                "label": "New type",
                "codes": [],
                "min_actors": 1,
            })
            st.rerun()

        if i < len(st.session_state.layer_defs) - 1:
            ld["max_to_next_km"] = st.slider(
                f"Maximum distance to stage {i+2} (km)",
                1, 300,
                int(ld.get("max_to_next_km", 30.0)),
                1,
                key=f"dist_{i}"
            )
        else:
            ld["max_to_next_km"] = None
            st.caption("Last stage : no distance to the next one.")

        st.divider()

    colA, colB = st.columns(2)
    with colA:
        run = st.button("Start detection")
    with colB:
        reset = st.button("Reset")


if reset:
    st.session_state.paths = []
    st.session_state.centroids_df = None
    st.session_state.selected_path = None
    st.session_state.eprtr_df = None
    st.session_state.uww_df = None
    st.session_state.layer_dfs = None
    st.session_state.pivot_selected_idx = 0
    st.session_state.export_html = None
    st.session_state.export_excel = None
    st.session_state.export_pdf = None
    st.session_state.export_zip = None
    st.session_state.export_artifacts = None
    st.rerun()


# ======================================================
# DETECTION
# ======================================================
if run:
    with st.spinner("Calculation of symbiotic networks"):
        eprtr = load_eprtr(eprtr_path)
        uww = load_uwwtps(uww_path)

        layer_dfs = []
        for ld in st.session_state.layer_defs:
            df_layer = filter_layer_df(ld, eprtr, uww)
            layer_dfs.append(df_layer)

        st.session_state.eprtr_df = eprtr
        st.session_state.uww_df = uww
        st.session_state.layer_dfs = layer_dfs

        for i, dfL in enumerate(layer_dfs):
            if dfL.empty:
                st.error(f"Stage {i+1} is empty after filtering : {st.session_state.layer_defs[i]['label']}")
                st.stop()

        dfP = layer_dfs[st.session_state.pivot_layer].copy()
        pivot_labels = _labels_for_layer(dfP)

        if st.session_state.pivot_mode == "Pivot stakeholder (selection)":
            if pivot_search.strip():
                mask = [pivot_search.lower() in s.lower() for s in pivot_labels]
                idx_map = [i for i, keep in enumerate(mask) if keep]
                labels_f = [pivot_labels[i] for i in idx_map]
            else:
                idx_map = list(range(len(pivot_labels)))
                labels_f = pivot_labels

            if len(labels_f) == 0:
                st.error("No pivot matches the filter.")
                st.stop()

            default_k = 0
            if st.session_state.pivot_selected_idx in idx_map:
                default_k = idx_map.index(st.session_state.pivot_selected_idx)

            chosen_k = st.sidebar.selectbox(
                f"Selection of the pivot (stage {st.session_state.pivot_layer+1})",
                options=list(range(len(labels_f))),
                index=min(max(default_k, 0), len(labels_f) - 1),
                format_func=lambda k: labels_f[k],
            )
            pivot_idx_real = int(idx_map[int(chosen_k)])
            st.session_state.pivot_selected_idx = pivot_idx_real
            pivot_indices = [pivot_idx_real]
        else:
            pivot_indices = list(range(len(dfP)))

        progress_bar = None
        if st.session_state.pivot_mode != "Pivot stakeholder (selection)" and len(pivot_indices) > 1:
            progress_bar = st.progress(0.0, text=f"Networks calculation : 0 / {len(pivot_indices)} treated pivots")

        def _update_progress(step, total):
            if progress_bar is not None and total > 0:
                progress_bar.progress(
                    min((step + 1) / total, 1.0),
                    text=f"Networks calculation : {step + 1} / {total} treated pivots"
                )

        paths = build_networks_bidirectional_pivot(
            st.session_state.layer_defs,
            layer_dfs,
            pivot_layer=st.session_state.pivot_layer,
            pivot_indices=pivot_indices,
            max_neighbors=max_neighbors,
            max_networks=max_paths,
            progress_callback=_update_progress if progress_bar is not None else None
        )

        if progress_bar is not None:
            progress_bar.empty()

        if len(paths) == 0:
            st.warning("No network found with these settings.")
            st.session_state.paths = []
            st.session_state.centroids_df = None
            st.session_state.selected_path = None
        else:
            st.session_state.paths = paths
            st.session_state.centroids_df = pd.DataFrame(
                [{"lat": p["lat"], "lon": p["lon"], "path_id": p["path_id"]} for p in paths]
            )
            st.session_state.selected_path = None

        st.session_state.heatmap_data = [
            [float(p["lat"]), float(p["lon"])]
            for p in paths
        ]

        st.session_state.export_html = None
        st.session_state.export_excel = None
        st.session_state.export_pdf = None
        st.session_state.export_zip = None
        st.session_state.export_artifacts = None


# ======================================================
# DISPLAY
# ======================================================
paths = st.session_state.paths
centroids_df = st.session_state.centroids_df
layer_defs = st.session_state.layer_defs
layer_dfs = st.session_state.layer_dfs

if not paths:
    st.info("Set your stages and pivot, then click 'Start detection'.")
    st.stop()

st.success(f"✅ {len(paths)} detected networks")

col_map, col_info = st.columns([2, 1], gap="large")


# ======================================================
# HISTOGRAMME : répartition des réseaux par pays (pivot)
# ======================================================
rows = []
for p in paths:
    piv_layer = int(p.get("pivot_layer", 0))
    piv_idx = int(p.get("pivot_idx", 0))

    if (
        layer_dfs is None
        or piv_layer >= len(layer_dfs)
        or piv_idx >= len(layer_dfs[piv_layer])
    ):
        continue

    row = layer_dfs[piv_layer].iloc[piv_idx]
    country = get_actor_country(row)
    rows.append({"path_id": int(p.get("path_id", -1)), "country": country})

df_net_countries = pd.DataFrame(rows)

if len(df_net_countries) == 0:
    counts = pd.Series(dtype=int)
else:
    counts = df_net_countries["country"].value_counts().sort_values(ascending=False)

fig_country = None
if len(counts) > 0:
    # Graphique interactif (zoom / survol) pour l'écran. L'export PDF
    # conserve son propre graphique matplotlib indépendant (cf.
    # build_pdf_bytes), qui n'est pas affecté par ce changement.
    df_counts_plot = counts.rename_axis("country").reset_index(name="count")
    fig_country = px.bar(
        df_counts_plot,
        x="country",
        y="count",
        color_discrete_sequence=["#02A717"],
        labels={"country": "Country", "count": "Number of potential replicated pilots"},
        title="Distribution of potential replicated pilots per country",
    )
    fig_country.update_traces(marker_line_color="black", marker_line_width=1)
    fig_country.update_layout(showlegend=False, margin=dict(t=50, b=40))


# ======================================================
# EXPORT PREPARATION
# ======================================================
artifacts = prepare_export_artifacts(
    paths=paths,
    layer_dfs=layer_dfs,
    layer_defs=layer_defs,
    counts=counts,
    pivot_layer=st.session_state.pivot_layer,
    max_neighbors=max_neighbors,
    max_paths=max_paths
)

with st.sidebar:
    st.divider()
    st.header("Sharing / exports")

    if st.button("Preparing partner exports"):
        with st.spinner("Preparing the files..."):
            html_bytes = build_overview_map_html(
                paths=paths,
                layer_dfs=layer_dfs,
                layer_defs=layer_defs,
                centroids_df=centroids_df,
                heatmap_data=st.session_state.heatmap_data,
                show_heatmap=show_heatmap
            )

            excel_bytes = build_excel_bytes(artifacts)

            pdf_bytes = build_pdf_bytes(
                centroids_df=centroids_df,
                counts=counts,
                artifacts=artifacts,
                paths=paths,
                layer_dfs=layer_dfs,
                layer_defs=layer_defs
            )

            zip_bytes = build_zip_bytes(
                html_bytes=html_bytes,
                excel_bytes=excel_bytes,
                pdf_bytes=pdf_bytes,
                artifacts=artifacts
            )

            st.session_state.export_html = html_bytes
            st.session_state.export_excel = excel_bytes
            st.session_state.export_pdf = pdf_bytes
            st.session_state.export_zip = zip_bytes
            st.session_state.export_artifacts = artifacts

    if st.session_state.export_html is not None:
        st.download_button(
            "Download the HTML map",
            data=st.session_state.export_html,
            file_name="replication_map.html",
            mime="text/html"
        )

    if st.session_state.export_excel is not None:
        st.download_button(
            "Download the Excel file",
            data=st.session_state.export_excel,
            file_name="replication_networks.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if st.session_state.export_pdf is not None:
        st.download_button(
            "Download the summary PDF",
            data=st.session_state.export_pdf,
            file_name="replication_summary.pdf",
            mime="application/pdf"
        )

    if st.session_state.export_zip is not None:
        st.download_button(
            "Download the full package (.zip)",
            data=st.session_state.export_zip,
            file_name="replication_package.zip",
            mime="application/zip"
        )


# ======================================================
# MAP
# ======================================================
with col_map:
    lat0 = float(centroids_df["lat"].mean())
    lon0 = float(centroids_df["lon"].mean())

    if st.session_state.selected_path is None:
        m = make_base_map(lat0, lon0, zoom=6)

        # Au-delà d'un grand nombre de réseaux, on regroupe les marqueurs
        # (MarkerCluster) pour garder la carte lisible et fluide ; en
        # dessous du seuil, comportement identique à avant (marqueurs directs).
        if len(centroids_df) > 150:
            marker_target = MarkerCluster(name="Detected networks").add_to(m)
        else:
            marker_target = m

        for _, r in centroids_df.iterrows():
            folium.CircleMarker(
                location=[r["lat"], r["lon"]],
                radius=4,
                color="#00B400",
                fill=True,
                fill_opacity=0.85,
                tooltip=f"Network {int(r['path_id'])}"
            ).add_to(marker_target)

        if show_heatmap and st.session_state.heatmap_data:
            HeatMap(
                st.session_state.heatmap_data,
                radius=35,
                blur=25,
                min_opacity=0.6,
                max_zoom=8,
            ).add_to(m)

        out = st_folium(
            m,
            height=740,
            use_container_width=True,
            returned_objects=["last_clicked"],
            key=f"main_map_{show_heatmap}"
        )

        st.markdown("### 📊 Distribution of networks by country")
        if fig_country is None:
            st.info("No country data is available to plot the histogram.")
        else:
            st.plotly_chart(fig_country, use_container_width=True)

        clicked = out.get("last_clicked")
        if clicked and "lat" in clicked and "lng" in clicked:
            pid = nearest_centroid_id(clicked["lat"], clicked["lng"], centroids_df, max_km=5.0)
            if pid is not None:
                st.session_state.selected_path = pid
                st.rerun()

    else:
        pid = int(st.session_state.selected_path)
        p = paths[pid]

        layer_sets = p["layer_sets"]
        layer_centroids = p.get("layer_centroids", [])

        m = make_base_map(p["lat"], p["lon"], zoom=8)

        def color_for_layer(li, src):
            if src == "UWWTPS":
                return "orange"
            if li == 0:
                return "red"
            if li == len(layer_sets) - 1:
                return "blue"
            return "purple"

        for li, idxs in enumerate(layer_sets):
            if idxs is None:
                continue
            dfL = layer_dfs[li]

            for ridx in idxs:
                row = dfL.iloc[int(ridx)]
                folium.CircleMarker(
                    [float(row.latitude), float(row.longitude)],
                    radius=5,
                    color=color_for_layer(li, row["__layer_source"]),
                    fill=True,
                    fill_opacity=0.85,
                    tooltip=f"{layer_defs[li]['label']} • {row['__layer_code']}",
                    popup=(
                        f"{layer_defs[li]['label']}<br>"
                        f"{row['__name']}<br>"
                        f"Code: {row['__layer_code']}<br>"
                        f"Source: {row['__layer_source']}"
                    )
                ).add_to(m)

        # LIAISONS acteur -> acteur conservées
        for li in range(len(layer_sets) - 1):
            idxs_src = layer_sets[li]
            idxs_tgt = layer_sets[li + 1]

            if not idxs_src or not idxs_tgt:
                continue

            df_src = layer_dfs[li]
            df_tgt_full = layer_dfs[li + 1]

            df_tgt = df_tgt_full.iloc[idxs_tgt].copy().reset_index(drop=True)

            tree_tgt = BallTree(
                np.radians(df_tgt[["latitude", "longitude"]].values),
                metric="haversine"
            )

            # Requête vectorisée (cf. build_links_table) au lieu d'une
            # requête BallTree par acteur dans la boucle de dessin.
            df_src_sel = df_src.iloc[idxs_src]
            src_coords = np.radians(df_src_sel[["latitude", "longitude"]].values)
            dist_all, ind_all = tree_tgt.query(src_coords, k=1)

            for row_i, ridx in enumerate(idxs_src):
                src_row = df_src_sel.iloc[row_i]
                src_lat = float(src_row.latitude)
                src_lon = float(src_row.longitude)

                tgt_row = df_tgt.iloc[int(ind_all[row_i][0])]

                line = folium.PolyLine(
                    locations=[
                        [src_lat, src_lon],
                        [float(tgt_row.latitude), float(tgt_row.longitude)]
                    ],
                    color="black",
                    weight=1.5,
                    opacity=0.6,
                ).add_to(m)

                try:
                    PolyLineTextPath(
                        line,
                        " ➤ ",
                        repeat=False,
                        center=True,
                        attributes={"font-size": "10"}
                    ).add_to(m)
                except Exception:
                    pass

        st_folium(m, height=740, use_container_width=True)


# ======================================================
# DETAILS
# ======================================================
with col_info:
    if st.session_state.selected_path is None:
        st.info("Click on a centroid to view the details of a network.")
        st.caption("Pivot: you choose the pivot stage + (optional) a pivot actor.")
    else:
        pid = int(st.session_state.selected_path)
        p = paths[pid]

        st.subheader(f"Network #{pid}")

        piv_layer = int(p.get("pivot_layer", st.session_state.pivot_layer))
        piv_idx = int(p.get("pivot_idx", 0))
        if layer_dfs is not None and piv_layer < len(layer_dfs) and piv_idx < len(layer_dfs[piv_layer]):
            piv = layer_dfs[piv_layer].iloc[piv_idx]
            st.markdown("### 🎯 Pivot")
            st.write({
                "etage": piv_layer + 1,
                "idx": piv_idx,
                "nom": piv.get("__name", ""),
                "code": piv.get("__layer_code", ""),
                "source": piv.get("__layer_source", ""),
                "pays": get_actor_country(piv),
                "lat": float(piv.latitude),
                "lon": float(piv.longitude),
            })

        st.markdown("### Stakeholders included by stage")
        rows = []
        for li, idxs in enumerate(p["layer_sets"]):
            if idxs is None:
                continue
            dfL = layer_dfs[li]
            for ridx in idxs:
                row = dfL.iloc[int(ridx)]
                rows.append({
                    "étage": li + 1,
                    "label": layer_defs[li]["label"],
                    "source": row["__layer_source"],
                    "code": row["__layer_code"],
                    "nom": row["__name"],
                    "pays": get_actor_country(row),
                    "lat": float(row.latitude),
                    "lon": float(row.longitude),
                })

        st.dataframe(pd.DataFrame(rows), use_container_width=True, height=520)

        if st.button("Back to centroids"):
            st.session_state.selected_path = None
            st.rerun()
